# -*- coding: utf-8 -*-

import argparse
import ast
import datetime
import json
import logging
import os
import re
import sys
#import urllib
import xml.dom.minidom
from xml.sax.saxutils import escape
from itertools import chain
#from urllib.parse import quote
#from bs4 import BeautifulSoup
from openpyxl import load_workbook

LOG_PATH = 'logs'
LOG_NAME = 'cmdline'
cross_check_identifier = re.compile('%.*?%')
logger = logging.getLogger()
cd = os.path.dirname(os.path.realpath(__file__))
ul_template = "\n     <gmd:useLimitation>\n" \
              "      <gmx:Anchor xlink:href=\"%%C_UL_U%%\">%%C_UL%%</gmx:Anchor>\n" \
              "     </gmd:useLimitation>\n"
ti_template = "\n    <gmd:title>\n" \
              "      <gco:CharacterString>%%MI_T%%</gco:CharacterString>\n" \
              "     </gmd:title>\n"
te_template = "\n     <gmd:description>\n" \
              "      <gmx:Anchor xlink:href=\"%%TE_DE_U%%\">%%TE_DE%%</gmx:Anchor>\n" \
              "     </gmd:description>\n"
ati_template = "\n  <gmd:alternateTitle>\n" \
               "      <gco:CharacterString>%%MI_AT%%</gco:CharacterString>\n" \
               "     </gmd:alternateTitle>\n"
lc_template = "\n   <gmd:otherConstraints >\n" \
              "     <gmx:Anchor xlink:href=\"%%C_LC_U%%\">%%C_LC%%</gmx:Anchor>\n" \
              "     </gmd:otherConstraints>\n"
kiso_template ="\n  <gmd:topicCategory>\n" \
               "    <gmd:MD_TopicCategoryCode>%%K_ISO%%</gmd:MD_TopicCategoryCode>\n" \
               "    </gmd:topicCategory>\n"
et_template = "\n   <gmd:keyword>\n" \
              "      <gmx:Anchor xlink:href=\"%%K_ET_U%%\">%%K_ET%%</gmx:Anchor>\n" \
              "     </gmd:keyword>\n"
et_template2 = "\n   <gmd:keyword>\n" \
              "     <gmx:Anchor xlink:href=\"https://gcmdservices.gsfc.nasa.gov/kms/concept/%%K_G_U%%\">%%K_G_C%% &gt; %%K_G_TO%% &gt; %%K_G_TE%% &gt; %%K_G_V%%</gmx:Anchor>\n" \
              "     </gmd:keyword>\n"
om_template = "   <gmd:descriptiveKeywords>\n" \
              "   <gmd:MD_Keywords>\n" \
              " %%K_O_VAL%%" \
              "    <gmd:type>\n" \
              "      <gmd:MD_KeywordTypeCode codeList=\"https://earth.esa.int/2017/resources/codeList.xml#MD_KeywordTypeCode\" codeListValue=\"%%K_O_VAR%%\"/>\n" \
              "    </gmd:type>\n" \
              "     <gmd:thesaurusName>\n" \
              "<gmd:CI_Citation>\n" \
              "<gmd:title>\n" \
              "<gmx:Anchor xlink:href=\"http://www.opengis.net/eop/2.1/\">Observations and Measurements Version 1.1</gmx:Anchor>\n" \
              "</gmd:title>\n" \
              "<gmd:date>\n" \
              "<gmd:CI_Date>\n" \
              "<gmd:date>\n" \
              "<gco:Date>2016-06-09</gco:Date>\n" \
              "</gmd:date>\n" \
              "<gmd:dateType>\n" \
              "<gmd:CI_DateTypeCode codeList=\"http://standards.iso.org/ittf/PubliclyAvailableStandards/ISO_19139_Schemas/resources/codelist/ML_gmxCodelists.xml#CI_DateTypeCode\" codeListValue=\"publication\">publication</gmd:CI_DateTypeCode>\n" \
              "</gmd:dateType>\n" \
              "</gmd:CI_Date>\n" \
              "</gmd:date>\n" \
              "</gmd:CI_Citation>\n" \
              "</gmd:thesaurusName>\n" \
              "</gmd:MD_Keywords>\n" \
              "</gmd:descriptiveKeywords>\n"
omval_template = "     <gmd:keyword>\n" \
              "         <gmx:Anchor xlink:href=\"%%K_O_U%%\">%%K_O_VAL%%</gmx:Anchor>\n" \
              "       </gmd:keyword>\n"
omeval_template = "     <gmd:keyword>\n" \
              "         <gmx:Anchor xlink:href=\"%%K_OE_U%%\">%%K_OE_VAL%%</gmx:Anchor>\n" \
              "       </gmd:keyword>\n"
ome_template = "<gmd:descriptiveKeywords>\n" \
               "    <gmd:MD_Keywords>\n" \
               " %%K_OE_VAL%%" \
              "     <gmd:type>\n" \
              "      <gmd:MD_KeywordTypeCode codeList=\"https://earth.esa.int/2017/resources/codeList.xml#MD_KeywordTypeCode\" codeListValue=\"%%K_OE_VAR%%\"/>\n" \
              "    </gmd:type>\n" \
              "     <gmd:thesaurusName>\n" \
              "    <gmd:CI_Citation>\n" \
               "       <gmd:title>\n" \
               "        <gmx:Anchor xlink:href=\"https://earth.esa.int/eop-ext/\">Observations and Measurements Extension</gmx:Anchor>\n" \
               "       </gmd:title>\n" \
               "       <gmd:date>\n" \
               "        <gmd:CI_Date>\n" \
               "         <gmd:date>\n" \
               "          <gco:Date>2017</gco:Date>\n" \
               "         </gmd:date>\n" \
               "         <gmd:dateType>\n" \
               "          <gmd:CI_DateTypeCode codeList=\"http://standards.iso.org/ittf/PubliclyAvailableStandards/ISO_19139_Schemas/resources/codelist/ML_gmxCodelists.xml#CI_DateTypeCode\" codeListValue=\"publication\">publication</gmd:CI_DateTypeCode>\n" \
               "         </gmd:dateType>\n" \
               "        </gmd:CI_Date>\n" \
               "       </gmd:date>\n" \
               "      </gmd:CI_Citation>\n" \
               "     </gmd:thesaurusName>\n" \
               "    </gmd:MD_Keywords>\n" \
               "   </gmd:descriptiveKeywords>\n"
fk_template = "<gmd:descriptiveKeywords>\n" \
              "    <gmd:MD_Keywords>\n" \
              "\n     <gmd:keyword>\n" \
              "         <gco:CharacterString>%%K_F%%</gco:CharacterString>\n" \
              "       </gmd:keyword>\n" \
              "    </gmd:MD_Keywords>\n" \
              "</gmd:descriptiveKeywords>\n"
kp_template = "<gmd:keyword>\n" \
              "  <gco:CharacterString>%%K_P%%</gco:CharacterString>\n" \
              "</gmd:keyword>\n"
mkp_template = "\n<gmd:descriptiveKeywords>\n" \
               "				<gmd:MD_Keywords>\n" \
               "					%%K_P%%" \
               "					<gmd:type>\n" \
               "						<gmd:MD_KeywordTypeCode codeList=\"http://standards.iso.org/ittf/PubliclyAvailableStandards/ISO_19139_Schemas/resources/codelist/gmxCodelists.xml#MD_KeywordTypeCode\" codeListValue=\"place\">place</gmd:MD_KeywordTypeCode>\n" \
               "					</gmd:type>\n" \
               "				</gmd:MD_Keywords>\n" \
               "</gmd:descriptiveKeywords>\n		"
pl_template = "\n   <gmd:contentInfo xmlns:gmd=\"http://www.isotc211.org/2005/gmd\">\n" \
              "         <gmi:MI_ImageDescription>\n" \
              "             <gmd:attributeDescription/>\n" \
              "                 <gmd:contentType/>\n" \
              "                     <gmd:processingLevelCode>\n" \
              "                         <gmd:RS_Identifier>\n" \
              "                             <gmd:code>\n" \
              "                                 <gco:CharacterString>%%PL%%</gco:CharacterString>\n" \
              "                             </gmd:code>\n" \
              "                         <gmd:codeSpace/>\n" \
              "                     </gmd:RS_Identifier>\n" \
              "                 </gmd:processingLevelCode>\n" \
              "         </gmi:MI_ImageDescription>\n" \
              "     </gmd:contentInfo>\n"
dt_template = "\n     <gmd:onLine>\n" \
              "        <gmd:CI_OnlineResource>\n" \
              "         <gmd:linkage>\n" \
              "          <gmd:URL>%%D_OR_U%%</gmd:URL>\n" \
              "         </gmd:linkage>\n" \
              "        <gmd:protocol>\n" \
              "         <gco:CharacterString>%%D_OR_P%%</gco:CharacterString>\n" \
              "        </gmd:protocol>\n" \
              "       <gmd:applicationProfile>\n" \
              "        <gco:CharacterString>%%D_OR_AP%%</gco:CharacterString>\n" \
              "       </gmd:applicationProfile>\n" \
              "       <gmd:name>\n" \
              "        <gco:CharacterString>%%D_OR_N%%</gco:CharacterString>\n" \
              "       </gmd:name>\n" \
              "       <gmd:description>\n" \
              "        <gco:CharacterString>%%D_OR_D%%</gco:CharacterString>\n" \
              "       </gmd:description>\n" \
              "       <gmd:function>\n" \
              "        <gmd:CI_OnLineFunctionCode codeList=\"http://standards.iso.org/ittf/PubliclyAvailableStandards/ISO_19139_Schemas/resources/codelist/ML_gmxCodelists.xml#CI_OnLineFunctionCode\" codeListValue=\"information\"/>\n" \
              "       </gmd:function>\n" \
              "      </gmd:CI_OnlineResource>\n" \
              "     </gmd:onLine>\n"
pf_template ="\n   <gmi:platform>\n" \
             "    <gmi:MI_Platform>\n" \
             "     <gmi:citation>\n" \
             "      <gmd:CI_Citation xmlns:gmd=\"http://www.isotc211.org/2005/gmd\">\n" \
             "       <gmd:title>\n" \
             "        <gmx:Anchor xlink:href=\"%%P_E_U%%\">%%P_E_I%%</gmx:Anchor>\n" \
             "       </gmd:title>\n" \
             "      <gmd:alternateTitle>\n" \
             "        <gmx:Anchor xlink:href=\"%%P_G_U%%\">%%P_G_SN%%</gmx:Anchor>\n" \
             "       </gmd:alternateTitle>\n" \
             "       <gmd:date>\n" \
             "        <gmd:CI_Date>\n" \
             "         <gmd:date>\n" \
             "          <gco:Date>%%P_E_LD%%</gco:Date>\n" \
             "         </gmd:date>\n" \
             "         <gmd:dateType>\n" \
             "          <gmd:CI_DateTypeCode codeList=\"http://www.isotc211.org/2005/resources/Codelist/gmxCodelists.xml#CI_DateTypeCode\" codeListValue=\"creation\"/>\n" \
             "         </gmd:dateType>\n" \
             "        </gmd:CI_Date>\n" \
             "       </gmd:date>\n" \
             "      </gmd:CI_Citation>\n" \
             "     </gmi:citation>\n" \
             "     <gmi:identifier>\n" \
             "      <gmd:MD_Identifier xmlns:gmd=\"http://www.isotc211.org/2005/gmd\">\n" \
             "       <gmd:code>\n" \
             "        <gmx:Anchor xlink:href=\"%%P_E_U%%\">%%P_E_I%%</gmx:Anchor>\n" \
             "       </gmd:code>\n" \
             "      </gmd:MD_Identifier>\n" \
             "     </gmi:identifier>\n" \
             "     <gmi:description>\n" \
             "      <gco:CharacterString>%%P_G_LN%%</gco:CharacterString>\n" \
             "     </gmi:description>\n" \
             "     <gmi:sponsor>\n" \
             "      <gmd:CI_ResponsibleParty xmlns:gmd=\"http://www.isotc211.org/2005/gmd\">\n" \
             "       <gmd:organisationName>\n" \
             "        <gco:CharacterString>%%P_E_O%%</gco:CharacterString>\n" \
             "       </gmd:organisationName>\n" \
             "       <gmd:role/>\n" \
             "      </gmd:CI_ResponsibleParty>\n" \
             "     </gmi:sponsor>\n" \
             "      %%I_%%" \
             "    </gmi:MI_Platform>\n" \
             "   </gmi:platform>\n"
ins_template ="\n   <gmi:instrument>\n" \
             "      <gmi:MI_Instrument>\n" \
             "       <gmi:citation>\n" \
             "        <gmd:CI_Citation xmlns:gmd=\"http://www.isotc211.org/2005/gmd\">\n" \
             "         <gmd:title>\n" \
             "          <gmx:Anchor xlink:href=\"%%I_E_U%%\">%%I_E_SN%%</gmx:Anchor>\n" \
             "         </gmd:title>\n" \
             "         <gmd:alternateTitle>\n" \
             "          <gmx:Anchor xlink:href=\"%%I_G_U%%\">%%I_G_SN%%</gmx:Anchor>\n" \
             "         </gmd:alternateTitle>\n" \
             "         <gmd:date>\n" \
             "          <gmd:CI_Date>\n" \
             "           <gmd:date>\n" \
             "            <gco:Date>%%P_E_LD%%</gco:Date>\n" \
             "           </gmd:date>\n" \
             "           <gmd:dateType>\n" \
             "            <gmd:CI_DateTypeCode codeList=\"http://www.isotc211.org/2005/resources/Codelist/gmxCodelists.xml#CI_DateTypeCode\" codeListValue=\"creation\"/>\n" \
             "           </gmd:dateType>\n" \
             "          </gmd:CI_Date>\n" \
             "         </gmd:date>\n" \
             "         <gmd:identifier>\n" \
             "          <gmd:MD_Identifier>\n" \
             "           <gmd:code>\n" \
             "            <gmx:Anchor xlink:href=\"%%I_E_U%%\">%%I_E_SN%%</gmx:Anchor>\n" \
             "           </gmd:code>\n" \
             "          </gmd:MD_Identifier>\n" \
             "         </gmd:identifier>\n" \
             "        </gmd:CI_Citation>\n" \
             "       </gmi:citation>\n" \
             "       <gmi:type>\n" \
             "        <gmi:MI_SensorTypeCode/>\n" \
             "       </gmi:type>\n" \
             "       <gmi:description>\n" \
             "        <gco:CharacterString>%%I_G_LN%%</gco:CharacterString>\n" \
             "       </gmi:description>\n" \
             "      </gmi:MI_Instrument>\n" \
             "     </gmi:instrument>\n"
rf_template = "     <gmd:onLine xlink:type=\"simple\" xlink:href=\"xpointer(%%D_OR_RF%%)\">\n      " \
              "         <gmd:CI_OnlineResource>\n" \
              "       <gmd:linkage>\n" \
              "        <gmd:URL>%%D_OR_U%%</gmd:URL>\n" \
              "       </gmd:linkage>\n" \
              "       <gmd:applicationProfile>\n" \
              "        <gco:CharacterString>%%D_OR_AP%%</gco:CharacterString>\n" \
              "       </gmd:applicationProfile>\n" \
              "       <gmd:name>\n" \
              "        <gco:CharacterString>%%D_OR_N%%</gco:CharacterString>\n" \
              "       </gmd:name>\n" \
              "       <gmd:description>\n" \
              "        <gco:CharacterString>%%D_OR_D%%</gco:CharacterString>\n" \
              "       </gmd:description>\n" \
              "       <gmd:function>\n" \
              "        <gmd:CI_OnLineFunctionCode codeList=\"http://standards.iso.org/ittf/PubliclyAvailableStandards/ISO_19139_Schemas/resources/codelist/ML_gmxCodelists.xml#CI_OnLineFunctionCode\" codeListValue=\"information\"/>\n" \
              "       </gmd:function>\n" \
              "      </gmd:CI_OnlineResource>\n" \
              "     </gmd:onLine>\n"
# mid_template = "<gmd:RS_Identifier>\n" \
#               "<gmd:code>\n" \
#               "<gco:CharacterString>%%MI_D%%</gco:CharacterString>\n" \
#               "</gmd:code>\n" \
#               "<gmd:codeSpace>\n" \
#               "<gco:CharacterString>http://doi.org</gco:CharacterString>\n" \
#               "</gmd:codeSpace>\n" \
#               "</gmd:RS_Identifier>\n"
mid_template = "<gmd:identifier>\n" \
            "<gmd:MD_Identifier>\n" \
            "    <gmd:authority>\n" \
            "        <gmd:CI_Citation>\n" \
            "            <gmd:title/>\n" \
            "            <gmd:date/>\n" \
            "            <gmd:citedResponsibleParty>\n" \
            "                <gmd:CI_ResponsibleParty>\n" \
            "                    <gmd:organisationName>\n" \
            "                        <gco:CharacterString>https://doi.org/</gco:CharacterString>\n" \
            "                    </gmd:organisationName>\n" \
            "                    <gmd:role>\n" \
            "                        <gmd:CI_RoleCode codeList=\"https://cdn.earthdata.nasa.gov/iso/resources/Codelist/gmxCodelists.xml#CI_RoleCode\" codeListValue=\"authority\">authority</gmd:CI_RoleCode>\n" \
            "                    </gmd:role>\n" \
            "                </gmd:CI_ResponsibleParty>\n" \
            "            </gmd:citedResponsibleParty>\n" \
            "        </gmd:CI_Citation>\n" \
            "    </gmd:authority>\n" \
            "    <gmd:code>\n" \
            "        <gco:CharacterString>%%MI_D%%</gco:CharacterString>\n" \
            "    </gmd:code>\n" \
            "    <gmd:codeSpace>\n" \
            "        <gco:CharacterString>https://doi.org</gco:CharacterString>\n" \
            "    </gmd:codeSpace>\n" \
            "    <gmd:description>\n" \
            "        <!-- NASA Earth Data requirement is that this string must contain \"DOI\"; no such requirement for ESA -->\n" \
            "        <gco:CharacterString>%%MI_D_C%%</gco:CharacterString>\n" \
            "    </gmd:description>\n" \
            "</gmd:MD_Identifier>\n" \
        "</gmd:identifier>\n" \

nins_template = "     <gmd:keyword>\n" \
                 "      <gmx:Anchor xlink:href=\"%%I_E_T_U%%\">%%I_E_T%%</gmx:Anchor>\n" \
                 "     </gmd:keyword>\n"

# Structure in template_list :  cond, template, rep, subrep, sublevel, subtemplate
template_list = [["Alternate title", ati_template, '%%MI_AT%%', '', '', ''],
                 ["Title", ti_template, '%%MI_T%%', '', '', ''],
                 ["Temporal extent", te_template, '%%TE_D%%', '', '', ''],
                 ['kp', mkp_template, '%%K_P%%', '%%K_P%%', 'location', kp_template],
                 ["Earth topics", et_template, '%%K_ET%%','%%K_G%%','',''],
                 ["Earth topics GCMD", et_template2, '%%K_G%%','%%K_ET%%','',''],
                 ["ISO topic category", kiso_template, '%%K_I%%','','',''],
                 ["Legal constraints", lc_template, '%%C_L%%','','',''],
                 ["Use limitations", ul_template, '%%C_U%%','','',''],
                 ["Observations and measurements", om_template, '%%K_O_%%', '%%K_O_VAL%%', 'Value', omval_template],
                 ["Observations and measurements extension", ome_template, '%%K_OE%%', '%%K_OE_VAL%%', 'Value',
                  omeval_template],
                 ["Free keywords", fk_template, '%%K_F%%','','',''],
                 ["Processing levels", pl_template, '%%PL%%','','',''],
                 ["Distributions", dt_template, '%%D_OR%%','%%D_OR_RF%%','',rf_template],
                 ["Mission info", pf_template, '%%P_%%','%%I_%%',"Instrument",ins_template],
                 ["DOI", mid_template, '%%MI_D%%', '', '', ''],
                 ["Instruments", nins_template, '%%I_E%%', '', '', '']]


def setup_cmd_args():
    """Setup command line arguments."""
    parser = argparse.ArgumentParser(description="Translate XLS files to appropriate XML format for ingestion  in FEDEO.", formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("xlsfile", help="The XLS file to parse")
    parser.add_argument('-outputdir', help="directory to output the XML file", default='output')
    parser.add_argument('-j', action='store_true', help="Also export JSON file")
    parser.add_argument('-p', action='store_true', help="Pretty print XML file")
    parser.add_argument('-o', action='store_true', help="Overwrite output XML file")
    parser.add_argument('-l', action='store_true', help="Skip I_G_LN as mandatory field", default=False)
    return parser.parse_args()


def setup_logging():
    # Default logging function
    log_format = logging.Formatter("%(asctime)s [%(levelname)-5.5s]  %(message)s")
    logger.setLevel(logging.INFO)

    if not os.path.exists(LOG_PATH):
        os.makedirs(LOG_PATH)

    file_handler = logging.FileHandler("{}/{}.log".format(LOG_PATH, LOG_NAME))
    file_handler.setFormatter(log_format)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(log_format)
    logger.addHandler(console_handler)


def check_all_green(worksheet, l=False):
    if not l:
        mandatory_fields = ["OI_ONS", "OI_ONL", "OI_PN", "OI__E", "MI_I", "MI_T", "MI_CD", "MI_UD", "MI_AB", "GE_W", "GE_E", "GE_S", "GE_N", "TE_SD", "C_UL", "K_ISO", "D_OR_N", "D_OR_U", "P_E_I", "P_E_LD", "P_G_SN", "I_E_U", "I_G_SN", "I_G_LN"]
    else:
        mandatory_fields = ["OI_ONS", "OI_ONL", "OI_PN", "OI__E", "MI_I", "MI_T", "MI_CD", "MI_UD", "MI_AB", "GE_W", "GE_E", "GE_S", "GE_N", "TE_SD", "C_UL", "K_ISO", "D_OR_N", "D_OR_U", "P_E_I", "P_E_LD", "P_G_SN", "I_E_U", "I_G_SN"]
    count = 0
    for row in range(1, 500):
        fieldcode = worksheet.cell(row=row, column=1).value
        fieldname = worksheet.cell(row=row, column=2).value
        fieldvalue = worksheet.cell(row=row, column=3).value
        for mf in mandatory_fields:
            if fieldcode == mf and fieldvalue is None:
                logger.error(
                    "Missing field value: code [{}], name [{}], value [{}]".format(fieldcode, fieldname, fieldvalue))
                count += 1
    if count > 0:
        return False
    else:
        return True


def get_type(cond, template, rep, subrep, sublevel, subtemplate):
    """Define template type according to the template list settings"""
    clean_rep = rep.replace("%","")
    clean_subrep = subrep.replace("%","")
    if subrep == '' and sublevel == '' and subtemplate == '':
        tetype = 1
    elif sublevel == 'location':
        tetype = 2
    elif sublevel == '' and subtemplate == '':
        tetype = 3
    elif sublevel == '':
        tetype = 4
    elif clean_rep == clean_subrep[:len(clean_rep)]:
        tetype = 5
    else:
        tetype = 6
    return tetype


def check_workbook_sheets(workbook):
    """
    Check all sheets in the workbook to obtain the list of valid sheets to further parse.
     Each sheet name should be the same as an existing .xml template (without the extension)
     :returns list of strings of the sheets to parse
    """
    sheets_in_file = workbook.sheetnames
    valid_sheets_template = []
    for sheet in sheets_in_file:
        # check if a template with that name exists in templates folder
        if os.path.isfile(os.path.join(cd,'templates', sheet + '.xml')):
            logger.info('{} is a valid template'.format(sheet))
            valid_sheets_template.append(sheet)
        else:
            logger.error('{} is not a valid template. Skipping'.format(sheet))
    return valid_sheets_template


def find_number_of_entries(sheet):
    current_row = 2
    current_cell = sheet.cell(row=current_row, column=1)
    number_of_entries = 0
    while current_cell.value is not None:
        number_of_entries += 1
        current_row += 1
        current_cell = sheet.cell(row=current_row, column=1)
    return number_of_entries


def multiple_replacer(string, replacements=[]):
    # replace multiple strings in string
    for h, r in replacements:
        string = string.replace(h, r)
    return string


def clean_field_val(val, fieldcode=""):
    # print(val,fieldcode)
    if isinstance(val, datetime.datetime):
        val = datetime.datetime.strptime(str(val),'%Y-%m-%d %H:%M:%S')
        val = val.strftime('%Y-%m-%d')
    # elif str(val).find("&")>0:
    #     val = quote(val, safe='')
    elif (fieldcode == "OI_PH" or fieldcode == "OI_F") and not val.find("+") >= 0:
        val = "+"+str(val)
    # elif str(val).find("µ") > 0:
    #     val = val.encode('utf-8').decode('cp1252')
    else:
        val = str(val)
    return escape(val.strip())


def pp_json(json_thing, sort=False, indents=4):
    # pretty print json dict
    if type(json_thing) is str:
        f = json.dumps(json.loads(json_thing), sort_keys=sort, indent=indents)
    else:
        f = json.dumps(json_thing, sort_keys=sort, indent=indents)
    return f


def get_list_in_list(list, loc):
    for l in list:
        if l[2] == loc:
            return l


if __name__ == '__main__':
    # Parse command line arguments
    args = setup_cmd_args()
    # Setup logging to stdout
    setup_logging()
    # Open the input excel
    wb = load_workbook(filename=args.xlsfile, data_only=True)
    # check valid sheets in excel and if not found return an error
    valid_sheets = check_workbook_sheets(wb)
    if len(valid_sheets) < 1:
        logger.error("No valid sheets were found. (sheet name should equal a template name to use). Exiting.")
        exit()
    logger.info("Parsing file {}".format(args.xlsfile))
    # Define input array data structures %Q%
    added_to_multilevel = []
    added_to_multilevel2 = []
    # Iterate on the sheets of the excel
    for sheet in valid_sheets:
        # Retrieve input main xml template from the subfolder templates
        maintemplate = os.path.join(cd, 'templates', sheet + '.xml')
        worksheet = wb[sheet]
        # Check manadatory fields are present in the input excel
        logger.info("Checking mandatory fields...")
        if not check_all_green(worksheet, args.l):
            logger.error("Mandatory fields in excel file are not all filed. Exiting...")
            exit()
        # Declare data empty dictionary that is dumped in json format
        data = {}
        json_data = json.dumps(data)
        # Iterate over template list declared at the beginning of the tool
        for cond, template, rep, subrep, sublevel, subtemplate in template_list:
            # Get template type according to settings of rep, subrep, sublevel, subtemplate
            tetype = get_type(cond, template, rep, subrep, sublevel, subtemplate)
            pf = {}
            pfc = 0
            mlc = 0
            platforms = cond+"#"
            pfnum = platforms + str(pfc)
            pf[pfnum] = {}
            if sublevel!="":
                instr = sublevel+"#"
                instrc = instr + str(mlc)
                pf[pfnum][sublevel] ={}
                pf[pfnum][sublevel][instrc] = {}
            done=0
            for row in range(1, 500):
                fieldcode=worksheet.cell(row=row, column=1).value
                fieldname=worksheet.cell(row=row, column=2).value
                fieldvalue=worksheet.cell(row=row, column=3).value
                if not (fieldcode is None and fieldname is None):
                    # print(rep, subrep, fieldcode, fieldname, fieldvalue)
                    clean_rep = rep.replace("%","")
                    if subrep!='':
                        clean_subrep = subrep.replace("%", "")
                        if (str(fieldcode)[:len(clean_rep)] == clean_rep) or (str(fieldcode)[:len(clean_subrep)] == clean_subrep):
                            if str(fieldcode)[:len(clean_subrep)] == clean_subrep and sublevel != "" and sublevel != "location":
                                try:
                                    pf[pfnum][sublevel][instrc][str(fieldcode)]
                                    mlc = mlc + 1
                                except:
                                    pass
                                if mlc != 0:
                                    instrc = instr + str(mlc)
                                    try:
                                        pf[pfnum][sublevel][instrc]
                                    except:
                                        pf[pfnum][sublevel][instrc] = {}
                                pf[pfnum][sublevel][instrc][str(fieldcode)] = str(clean_field_val(fieldvalue, fieldcode))
                                done = 1
                                added_to_multilevel.append(str(fieldcode))
                            elif str(fieldcode)[:len(clean_rep)] == clean_rep and sublevel != "location":
                                try:
                                    pf[pfnum][str(fieldcode)]
                                    pfc = pfc + 1
                                except:
                                    pass
                                if pfc != 0:
                                    pfnum = platforms + str(pfc)
                                    try:
                                        pf[pfnum]
                                    except:
                                        pf[pfnum] = {}
                                if sublevel!="":
                                    try:
                                        pf[pfnum][sublevel][instrc]
                                    except:
                                        pf[pfnum][sublevel] = {}
                                        pf[pfnum][sublevel][instrc] = {}
                                pf[pfnum][str(fieldcode)] = str(clean_field_val(fieldvalue, fieldcode))
                                done=1
                                added_to_multilevel.append(str(fieldcode))
                            elif str(fieldcode)[:len(clean_rep)] == clean_rep and sublevel == "location" and not fieldvalue is None:
                                for split in fieldvalue.split(","):
                                    try:
                                        pf[pfnum][sublevel][instrc][str(fieldcode)]
                                        mlc = mlc + 1
                                    except:
                                        pass
                                    if mlc != 0:
                                        instrc = instr + str(mlc)
                                        try:
                                            pf[pfnum][sublevel][instrc]
                                        except:
                                            pf[pfnum][sublevel][instrc] = {}
                                    pf[pfnum][sublevel][instrc][str(fieldcode)] = str(clean_field_val(split, fieldcode))
                                    done = 1
                                    added_to_multilevel.append(str(fieldcode))
                    elif cond != '':
                        if str(fieldcode)[:len(clean_rep)] == clean_rep:
                            try:
                                pf[pfnum][str(fieldcode)]
                                pfc = pfc + 1
                            except:
                                pass
                            if pfc != 0:
                                pfnum = platforms + str(pfc)
                                try:
                                    pf[pfnum]
                                except:
                                    pf[pfnum] = {}
                            pf[pfnum][str(fieldcode)] = str(clean_field_val(fieldvalue, fieldcode))
                            if (cond == 'DOI') and (fieldcode == 'MI_D_C'):
                                if ('MI_D' not in pf[pfnum]) or (pf[pfnum]['MI_D'] == 'None'):
                                    pf[pfnum] = {}
                            done=1
                            added_to_multilevel.append(str(fieldcode))
                        elif not ("%%"+str(fieldcode)[:4]+"%%" in chain.from_iterable(template_list)) and not ("%%"+str(fieldcode)[:3]+"%%" in chain.from_iterable(template_list)) and not ("%%"+str(fieldcode)[:2]+"%%" in chain.from_iterable(template_list)) and not str(fieldcode) in added_to_multilevel:
                            data[str(fieldcode)] = str(clean_field_val(fieldvalue, fieldcode))
                            json_data = json.dumps(data)
                if cond == "" and not str(fieldcode) in added_to_multilevel and not fieldvalue is None:
                    data[str(fieldcode)] = str(clean_field_val(fieldvalue, fieldcode))
                    json_data = json.dumps(data)
            if done==1:
                pfc = '{\''+cond+'\':'+str(pf)+'}'
                pfc = ast.literal_eval(pfc)
                data.update(pfc)
                json_data = json.dumps(data)
        j = json.loads(json_data)
        complete_xml_path = os.path.join(args.outputdir, os.path.basename(args.xlsfile).split(".")[0]+".xml")
        complete_json_path = os.path.join(args.outputdir, os.path.basename(args.xlsfile).split(".")[0]+".json")
        # Create outputdir
        try:
            os.makedirs(os.path.dirname(args.outputdir))
        except OSError:
            pass
        # read main template from file
        with open(maintemplate, 'r') as file:
            filedata = file.read()
        nfiledata = filedata
        # replace all occurences of %%LOCATORS%% in template with key/values from json dict. When there are nested elements, recur to respective sub-templates and include them as many times as needed.
        for i in j:
            for cond, template, rep, subrep, sublevel, subtemplate in template_list:
                tetype = get_type(cond, template, rep, subrep, sublevel, subtemplate)
                if i == cond:
                    n_template = ""
                    u_template = ""
                    s_template = ""
                    m_template = ""
                    o_template = ""
                    for h in j[i]:
                        cross_check_value = ""
                        l = []
                        s = []
                        for x in j[i][h]:
                            if x == sublevel:
                                if tetype == 5:
                                    u_template = ""
                                for y in j[i][h][x]:
                                    g = []
                                    if len(y) < 2:
                                        continue
                                    for b in j[i][h][x][y]:
                                        if not (j[i][h][x][y][b]== 'None' or j[i][h][x][y][b] == "#N/A"):
                                            u = ['%%' + b + '%%', j[i][h][x][y][b]]
                                            if not u in g:
                                                g.append(u)
                                    if not j[i][h][x][y] == {}:
                                        if tetype == 5:
                                            # print("check0", j[i][h][x][y])
                                            if not j[i][h][x][y][subrep.replace("%","")] == 'None':
                                                u_template = u_template + multiple_replacer(subtemplate, g)
                                        else:
                                            u_template = u_template + multiple_replacer(subtemplate, g)
                            else:
                                # the particular case of K_G_U key is to be excluded if equals 0
                                if x == "K_G_U":
                                    if not (j[i][h][x] == 'None' or j[i][h][x] == "#N/A" or j[i][h]["K_G_U"] == "0"):
                                        v = ['%%' + x + '%%', j[i][h][x]]
                                        if not v in l:
                                            l.append(v)
                                elif not (j[i][h][x] == 'None' or j[i][h][x] == "#N/A"):
                                    v = ['%%'+x+'%%', j[i][h][x]]
                                    if not v in l:
                                        l.append(v)
                                else:
                                    # write white cross_check value. This is used in conditional cases where a template is to be re-writen conditionally with "subtemplate-key" value. For example, use identifier x if y is null.
                                    if cross_check_identifier.match(subtemplate):
                                        template = template.replace(rep,"%"+subtemplate+"%")
                                        v = ["%"+subtemplate+"%", j[subtemplate.replace("%","")]]
                                        if not v in l:
                                            l.append(v)

                        if subrep.replace("%","") in j[i][h] and j[i][h][subrep.replace("%","")] != 'None':
                            s_template = s_template + multiple_replacer(subtemplate, l)
                        elif tetype == 5:
                            # the particular cases of K_O_VAL and K_OE_VAL key, main template is not writen if no valid subtemplate was generated
                            if (rep == "%%K_O_%%" or rep == "%%K_OE%%") and u_template == "":
                                pass
                            else:
                                n_template = n_template + multiple_replacer(template, l)
                                n_template = n_template.replace(subrep, u_template)
                                n_template = multiple_replacer(n_template, l)
                        elif sublevel != "" and tetype != 5:
                            # print("check", rep, subrep, sublevel)
                            u_template = multiple_replacer(u_template, l)
                            m_template = template.replace(subrep, u_template)
                            m_template = multiple_replacer(m_template, l)
                            u_template = ""
                            o_template = o_template + m_template
                        else:
                            if len(l) > 0:
                                n_template = n_template + multiple_replacer(template, l)
                            elif (len(l) == 0) and (rep == '%%MI_D%%'):
                                n_template = '<gmd:identifier />'
                    if subrep != "" and subtemplate != "" and sublevel == "":
                        nfiledata = nfiledata.replace(subrep, s_template)
                    if o_template != "":
                        nfiledata = nfiledata.replace(rep, o_template)
                    nfiledata = nfiledata.replace(rep, n_template)
            if not (i in chain.from_iterable(template_list)) and not (j[i] == 'None' or j[i] == "#N/A"):
                nfiledata = nfiledata.replace('%%' + str(i) + '%%', str(j[i]))
        # Clean all unused %%LOCATORS%% in template
        nfiledata = re.sub('%%MI_D%%', '<gmd:identifier />', nfiledata)
        nfiledata = re.sub('%%.*?%%', '', nfiledata)
        nfiledata = nfiledata.replace("<gco:Date></gco:Date>", "")
        nfiledata = nfiledata.replace("> &gt; ", ">  &gt; ").replace("  &gt; ", "").replace(" &gt; <", "<").replace(" &gt;<", "<")
        # Decode urls
        urls = re.findall('"http.?.*?"', nfiledata)
        for url in urls:
            if url.find("gmd")<=0 and url.find("3.2")<=0 and url.find("isotc211")<=0 and url.find("w3.org")<=0:
                # encodedurl = quote(url.replace('\"',''), safe='')
                # nfiledata = nfiledata.replace(url, '\"' + encodedurl + '\"')
                # cleanurl = urllib.parse.unquote(url)
                encodedurl = url.replace('&','&amp;')
                nfiledata = nfiledata.replace(url, encodedurl)
        nfiledata = re.sub('&(?!amp;|gt;|lt;)', '&amp;', nfiledata)
        # Remove specific empty tags
        # tags = ['gmi:platform','gmi:instrument']
        # for tag in tags:
        #     maintags = re.findall(
        #         '<'+tag+'>.*?</'+tag+'>',
        #         nfiledata, re.DOTALL)
        #     for maintag in maintags:
        #         check = 0
        #         emptytags = re.findall(
        #             '<gmi:description>.*?<gco:CharacterString></gco:CharacterString>.*?</gmi:description>',
        #             maintag, re.DOTALL)
        #         for emptytag in emptytags:
        #             print(emptytag)
        #             newmaintag = maintag.replace(emptytag,'<gmi:description></gmi:description>')
        #             check = 1
        #         if check == 1:
        #             nfiledata = nfiledata.replace(maintag, newmaintag)
        # Pretty print xml output
        if args.p:
            xml = xml.dom.minidom.parseString(nfiledata.replace("\n"," "))
            nfiledata = xml.toprettyxml()
            # bs = BeautifulSoup(nfiledata, 'xml')
            # nfiledata = bs.prettify()
        # Write the file out again
        nfiledata = nfiledata.replace("<?xml version=\"1.0\" ?>", "<?xml version=\"1.0\" encoding=\"UTF-8\" ?>")
        with open(complete_xml_path, 'w', encoding="utf-8") as file:
            file.write(nfiledata)
        # Also export json file
        if args.j:
            with open(complete_json_path, 'w', encoding="utf-8") as file:
                file.write(str(pp_json(j)))
    logger.info("Done")
